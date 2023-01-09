import socket
from netaddr import IPRange
from openpyxl import Workbook
from datetime import datetime, date
from ipwhois import IPWhois
import os
from progressbar import ProgressBar

wb = Workbook()

ws=wb.active

ws['A1'] = "IP Address"
ws['B1'] = "DNS Name"
ws['B1'] = "CIDR"
ws['C1'] = "Country"
ws['D1'] = "Description"
ws['E1'] = "Date"

pbar = ProgressBar()
with open('inputFile.txt') as f:
    for line in pbar(f):
        if "-" in line:
                start_address = line.split("-")[0]
                end_address = line.split("-")[1]
                Range = IPRange(f'{start_address}', f'{end_address}') 
                for ip in Range:
                    DNSname = str(socket.getnameinfo((f'{ip}', 0), 0)).split("'")[1]
                    data = IPWhois(f'{ip}')
                    results = data.lookup_whois()
                    cidr = results["asn_cidr"]
                    country = results["asn_country_code"]
                    description = results["asn_description"]
                    date2 = results['asn_date']
                    print(ip, DNSname, cidr, country, description, date2)
                    ip = str(ip)
                    ws.append([ip, DNSname, cidr, country, description, date2])
        else:
                ip = line.strip("\n")
                DNSname = str(socket.getnameinfo((ip, 0), 0)).split("'")[1]
                data = IPWhois(f'{ip}')
                results = data.lookup_whois()
                cidr = results["asn_cidr"]
                country = results["asn_country_code"]
                description = results["asn_description"]
                date2 = results['asn_date']
                print(ip, DNSname, cidr, country, description, date2)
                ws.append([ip, DNSname, cidr, country, description, date2])

print("----------------------------")
print("You are DONE SIR!!! ( ͡° ͜ʖ ͡°)")
print("----------------------------")

currenttime = datetime.now()
timestr = currenttime.strftime("%H%M%S")
Current_Directory = os.getcwd()

wb.save(Current_Directory+"/Results/External_IP_Investigation" + "_" + str(date.today()) + "_" + str(timestr) + ".xlsx")