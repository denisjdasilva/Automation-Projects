from ipwhois.net import Net
from ipwhois.asn import IPASN
from openpyxl import Workbook
from datetime import datetime, date
from progressbar import ProgressBar
import os

currenttime = datetime.now()
timestr = currenttime.strftime("%H%M%S")

wb = Workbook()

ws=wb.active

ws['A1'] = "IP Address"
ws['B1'] = "CIDR"
ws['C1'] = "Country"
ws['D1'] = "Description"
ws['E1'] = "Date"

count = 0 
pbar = ProgressBar()
with open('inputFile.txt') as f:
        for ip in pbar(f):
                    ip = str(ip.strip("\n"))
                    net = Net(f'{ip}')
                    data = IPASN(net)
                    results = data.lookup()
                    cidr = results["asn_cidr"]
                    country = results["asn_country_code"]
                    description = results["asn_description"]
                    date2 = results['asn_date']
                    ws.append([ip, cidr, country, description, date2])
                    #print(ip, cidr, country, description, date2)

                    count += 1
                    if count < 1000:
                        continue
                    else:
                        count = 0
                        wb.save("WhoisLookup_TEMP_Results" + "_" + str(date.today()) + "_" + str(timestr) + ".xlsx")

print("----------------------------")
print("You are DONE SIR!!! ( ͡° ͜ʖ ͡°)")
print("----------------------------")

currenttime = datetime.now()
timestr = currenttime.strftime("%H%M%S")
Current_Directory = os.getcwd()
wb.save(Current_Directory+"/Results/WhoisLookup_Results" + "_" + str(date.today()) + "_" + str(timestr) + ".xlsx")


###PROTOCOLS BELOW FOR TESTING###

# IP = '200.14.16.1'

# # #IP ASN, port 80, fastest
# # net = Net(f'{IP}')
# # obj = IPASN(net)
# # results = obj.lookup('asn_methods=dns')
# # print(results)

# # #Legacy, IPWhois, port 43
# # obj = IPWhois(f'{IP}')
# # results = obj.lookup_whois()
# # print(results)

# #RDAP, slowest, most data
# obj = IPWhois(f'{IP}')
# results = obj.lookup_rdap('asn_methods=http')
# cidr = results['network']['handle']
# registrant = results['objects']['MX-THMS-LACNIC']['contact']['name']
# email = str(results['objects']['MAR27']['contact']['email'])
# contact = email.split("'")[5]
# print(contact)