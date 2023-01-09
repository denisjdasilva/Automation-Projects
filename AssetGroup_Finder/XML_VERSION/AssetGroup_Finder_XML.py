from bs4 import BeautifulSoup
from netaddr import IPRange
from openpyxl import Workbook
from datetime import datetime, date
from progressbar import ProgressBar
import os

wb = Workbook()

ws=wb.active

ws['A1'] = "IP ADDRESS"
ws['B1'] = "ASSET GROUP"

with open('AG_inputFile.xml') as AG_f:
    data = AG_f.read()
data = BeautifulSoup(data, 'xml')

pbar = ProgressBar()
with open('IP_inputFile.txt') as IP_f:
    for IP in pbar(IP_f):
        for record in data.find_all('RECORD'):
            title = str(record.find_all('KEY', {'name':'TITLE'}))
            title = title.split(">")[1]
            title = title.split("<")[0]
            ips = str(record.find_all('KEY', {'name':'IPS'}))
            ips = ips.split(">")[1]
            ips = ips.split("<")[0]
            RangeSet = ips.split(",")
            for IPset in RangeSet:
                    if "-" in IPset:
                        start_address = IPset.split("-")[0]
                        end_address = IPset.split("-")[1]
                        range = IPRange(f'{start_address}', f'{end_address}')
                        if IP in range:
                            ws.append([IP, title])
                        else:   
                            continue
                    else:
                        if IP == IPset:
                            ws.append([IP, title])
                        else:
                            continue

print("----------------------------")
print("You are DONE SIR!!! ( ͡° ͜ʖ ͡°)")
print("----------------------------")

currenttime = datetime.now()
timestr = currenttime.strftime("%H%M%S")
Current_Directory = os.getcwd()
wb.save(Current_Directory+'/Results/AssetGroupFinder_Results' + "_" + str(date.today()) + "_" + str(timestr) + ".xlsx")
